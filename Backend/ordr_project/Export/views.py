from django.shortcuts import render

# Create your views here.
import os
import pandas as pd
from datetime import datetime
import calendar
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl.styles import Alignment, Font
from .models import ExcelFile
from .serializers import ExcelFileSerializer
from django.conf import settings

class GenerateORDRSheetView(APIView):
    def post(self, request, *args, **kwargs):
        # Serialize and validate the uploaded file
        serializer = ExcelFileSerializer(data=request.data)
        if serializer.is_valid():
            excel_file_instance = serializer.save()
            file_path = excel_file_instance.file.path

            try:
                # Automatically detect the first sheet in the file
                df_spd = pd.read_excel(file_path, sheet_name=0)  # Automatically uses the first sheet

                # Mapping the existing columns to the required columns
                ordr_column_mapping = {
                    'DocNum': 'CustomerRefNo',
                    'DocType': 'ItemDescription',
                    'Series': 'Series',
                    'NumAtCard': 'CustomerRefNo',
                    'DocDate': 'Document Date',
                    'TaxDate': 'Tax Date',
                    'DocDueDate': 'Document Due Date',
                    'DocCur':     'DocCur',
                    'DocRate':'Docrate',
                    'CardCode': 'Customer Code',
                    'U_FRIEGHT': 'Frieght',
                    'U_SALESCAT': 'Sales Category',
                    'U_DEPT': 'Department',
                    'U_MHXML': 'Part No',
                }

                # Create a DataFrame for the ORDR sheet with required columns
                df_ordr_required = df_spd[list(ordr_column_mapping.values())].rename(columns=ordr_column_mapping)

                # Adjust the column names for final output
                df_ordr_required.columns = ['DocNum', 'DocType', 'Series', 'NumAtCard', 'DocDate', 'TaxDate', 'DocDueDate', 'CardCode', 
                                            'U_FRIEGHT', 'U_SALESCAT', 'U_DEPT', 'U_MHXML']

                # Set the 'U_XmlFileStatus' column to 1 by default
                df_ordr_required['U_XmlFileStatus'] = 1

                # Insert the 'BPL_IDAssignedToInvoice' column with a default value of 3
                df_ordr_required['BPL_IDAssignedToInvoice'] = 3

                # Adjust the columns according to the new order specified
                new_columns_order = ['Series', 'NumAtCard', 'DocDate', 'TaxDate', 'DocDueDate', 'CardCode', 
                                     'U_FRIEGHT', 'U_SALESCAT', 'U_DEPT', 'U_XmlFileStatus', 'BPL_IDAssignedToInvoice', 
                                     'U_MHXML']
                
                # Re-arrange the DataFrame according to the new column order
                df_final_ordr_reordered_new = df_ordr_required[new_columns_order]
                
                # Insert 'DocNum' and 'DocType' columns with default values
                df_final_ordr_reordered_new.insert(0, 'DocNum', range(1, len(df_final_ordr_reordered_new) + 1))  # Incremental numbers
                df_final_ordr_reordered_new.insert(1, 'DocType', 'dDocument_Items')  # Default value for DocType

                # Set the 'Series' column to a default value of 881
                df_final_ordr_reordered_new['Series'] = 881

                # Set the 'DocDate' and 'TaxDate' columns to the current system date without separators (format YYYYMMDD)
                current_date = datetime.now().strftime('%Y%m%d')  # Format the date as YYYYMMDD
                df_final_ordr_reordered_new['DocDate'] = current_date
                df_final_ordr_reordered_new['TaxDate'] = current_date

                # Calculate the last day of the current month for 'DocDueDate'
                now = datetime.now()
                last_day = calendar.monthrange(now.year, now.month)[1]  # Get the last day of the current month
                last_day_of_month = now.replace(day=last_day).strftime('%Y%m%d')  # Format as YYYYMMDD
                df_final_ordr_reordered_new['DocDueDate'] = last_day_of_month

                # Set 'U_MHXML' based on the condition of 'CardCode'
                df_final_ordr_reordered_new['U_MHXML'] = df_final_ordr_reordered_new['CardCode'].apply(
                    lambda x: '0' if '16M' in x or '16G' in x else '1'
                )

                # Duplicate the column headers as the first row
                headers_as_row = pd.DataFrame([df_final_ordr_reordered_new.columns], columns=df_final_ordr_reordered_new.columns)
                headers_as_row.columns = df_final_ordr_reordered_new.columns  # Ensure same column names

                # Concatenate the duplicated header with the actual data
                df_final_ordr_reordered_new = pd.concat([headers_as_row, df_final_ordr_reordered_new], ignore_index=True)

                # In the first row, specifically change the value of 'BPL_IDAssignedToInvoice' to 'BPLId'
                df_final_ordr_reordered_new.at[0, 'BPL_IDAssignedToInvoice'] = 'BPLId'

                # Save the final ORDR sheet as Excel
                output_filename_xlsx = 'Generated_ORDR_Sheet.xlsx'
                output_path_xlsx = os.path.join(settings.MEDIA_ROOT, output_filename_xlsx)
                df_final_ordr_reordered_new.to_excel(output_path_xlsx, index=False)

                # Open the Excel file to apply center alignment to all content
                wb = openpyxl.load_workbook(output_path_xlsx)
                ws = wb.active

                # Apply center alignment to all cells
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                wb.save(output_path_xlsx)

                # Save the final ORDR sheet as TXT
                output_filename_txt = 'Generated_ORDR_Sheet.txt'
                output_path_txt = os.path.join(settings.MEDIA_ROOT, output_filename_txt)
                df_final_ordr_reordered_new.to_csv(output_path_txt, index=False, sep='\t')

                # Return the relative URL of both files
                file_url_xlsx = f'{settings.MEDIA_URL}{output_filename_xlsx}'
                file_url_txt = f'{settings.MEDIA_URL}{output_filename_txt}'
                return Response({"message": "ORDR sheet generated successfully", 
                                 "file_path_xlsx": file_url_xlsx, 
                                 "file_path_txt": file_url_txt}, status=status.HTTP_200_OK)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from .models import ExcelFile
from .serializers import ExcelFileSerializer
from django.conf import settings

class GenerateRDR1SheetView(APIView):
    def post(self, request, *args, **kwargs):
        serializer = ExcelFileSerializer(data=request.data)
        if serializer.is_valid():
            excel_file_instance = serializer.save()
            file_path = excel_file_instance.file.path

            try:
                # Automatically detect the first sheet in the file
                df_spd = pd.read_excel(file_path, sheet_name=0)  # Automatically uses the first sheet

                # Mapping the existing columns to the required columns for RDR1
                rdr1_column_mapping = {
                    'ItemCode': 'Item Code',
                    'SubCatNum': 'Part No',
                    'Quantity': 'Quantity',
                    'PriceBefDi': 'Price',
                    'TaxCode': 'TaxCode',
                    'WhsCode': 'Warehouse',
                }

                # Create a DataFrame for the RDR1 sheet with required columns
                df_rdr1_required = df_spd[list(rdr1_column_mapping.values())].rename(columns=rdr1_column_mapping)

                # Insert 'DocNum' as an incremental number starting from 1
                df_rdr1_required.insert(0, 'DocNum', range(1, len(df_rdr1_required) + 1))

                # Insert 'LineNum' as 0 by default for all rows
                df_rdr1_required.insert(1, 'LineNum', 0)

                # Adjust the column names for final output
                df_rdr1_required.columns = ['DocNum', 'LineNum', 'ItemCode', 'SubCatNum', 'Quantity', 'PriceBefDi', 'TaxCode', 'WhsCode']

                # Add the two rows manually at the start of the DataFrame
                first_row = pd.DataFrame([['ParentKey', 'LineNum', 'ItemCode', 'SupplierCatNum', 'Quantity', 'UnitPrice', 'TaxCode', 'WarehouseCode']],
                                         columns=df_rdr1_required.columns)
                second_row = pd.DataFrame([['DocNum', 'LineNum', 'ItemCode', 'SubCatNum', 'Quantity', 'PriceBefDi', 'TaxCode', 'WhsCode']],
                                          columns=df_rdr1_required.columns)

                # Prepend the two rows to the DataFrame
                df_rdr1_final = pd.concat([first_row, second_row, df_rdr1_required], ignore_index=True)

                # Save the final RDR1 sheet as Excel without headers
                output_filename_xlsx = 'Generated_RDR1_Sheet.xlsx'
                output_path_xlsx = os.path.join(settings.MEDIA_ROOT, output_filename_xlsx)
                df_rdr1_final.to_excel(output_path_xlsx, index=False, header=False)

                # Open the saved Excel file and apply bold formatting to the first row and center alignment to all content
                wb = openpyxl.load_workbook(output_path_xlsx)
                ws = wb.active

                # Make the first row bold and apply center alignment
                for cell in ws[1]:  # Loop through the first row
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                # Apply center alignment to all other cells
                for row in ws.iter_rows(min_row=2):  # Start from the second row
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                # Save the changes back to the Excel file
                wb.save(output_path_xlsx)

                # Save the final RDR1 sheet as TXT without headers
                output_filename_txt = 'Generated_RDR1_Sheet.txt'
                output_path_txt = os.path.join(settings.MEDIA_ROOT, output_filename_txt)
                df_rdr1_final.to_csv(output_path_txt, index=False, sep='\t', header=False)

                # Return the relative URL of both files
                file_url_xlsx = f'{settings.MEDIA_URL}{output_filename_xlsx}'
                file_url_txt = f'{settings.MEDIA_URL}{output_filename_txt}'
                return Response({"message": "RDR1 sheet generated successfully", 
                                 "file_path_xlsx": file_url_xlsx, 
                                 "file_path_txt": file_url_txt}, status=status.HTTP_200_OK)

            except Exception as e:
                return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
